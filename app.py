import streamlit as st
import pandas as pd
import openpyxl
import re
from io import BytesIO
from datetime import date, datetime, timedelta
from openpyxl.styles import PatternFill

# ============================================================
# アプリ設定（最初に）
# ============================================================
st.set_page_config(page_title="交通費CSV→テンプレ増分転記", layout="wide")

# UIちょい改善（ドロップ枠を少し広めに）
st.markdown("""
<style>
[data-testid="stFileUploader"] section {
  padding: 18px;
}
</style>
""", unsafe_allow_html=True)

# ============================================================
# ユーティリティ
# ============================================================

def read_csv_safely(uploaded_file) -> pd.DataFrame:
    raw = uploaded_file.getvalue()  # read()より安定
    for enc in ("utf-8-sig", "cp932", "utf-8"):
        try:
            return pd.read_csv(BytesIO(raw), encoding=enc)
        except Exception:
            continue
    raise RuntimeError("CSVが読み込めませんでした（文字コード/形式を確認）")

def find_col(df, candidates):
    for c in candidates:
        if c in df.columns:
            return c
    return None

def excel_serial_to_date(n: int) -> date:
    return date(1899, 12, 30) + timedelta(days=int(n))

def get_date_from_cell(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and value > 30000:
        return excel_serial_to_date(int(value))
    if isinstance(value, str):
        s = value.strip()
        m = re.search(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", s)
        if m:
            y, mo, d = map(int, m.groups())
            return date(y, mo, d)
        m = re.search(r"(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日", s)
        if m:
            y, mo, d = map(int, m.groups())
            return date(y, mo, d)
    return None

def find_header_row(ws, header_text="日付", search_max_row=120) -> int | None:
    for r in range(1, search_max_row + 1):
        for c in range(1, 40):
            if ws.cell(r, c).value == header_text:
                return r
    return None

def norm_str(v) -> str:
    return str(v).strip() if v is not None else ""

def norm_int(v) -> int:
    try:
        return int(pd.to_numeric(v, errors="coerce") or 0)
    except Exception:
        return 0

def collect_existing_keys(ws, start_row: int, max_scan: int = 3000):
    """
    テンプレに既に入っている行を読み取り、重複判定用キーのsetを返す
    キー: (日付, 訪問先・目的地, 交通手段, 移動区間, 金額)
    """
    keys = set()
    for r in range(start_row, start_row + max_scan):
        v_date = ws.cell(r, 1).value
        v_dest = ws.cell(r, 2).value
        v_mode = ws.cell(r, 3).value
        v_route = ws.cell(r, 4).value
        v_amt = ws.cell(r, 5).value

        if all(v is None for v in [v_date, v_dest, v_mode, v_route, v_amt]):
            break

        dd = get_date_from_cell(v_date)
        if dd is None:
            continue

        key = (dd, norm_str(v_dest), norm_str(v_mode), norm_str(v_route), norm_int(v_amt))
        keys.add(key)

    return keys

def find_first_empty_row(ws, start_row: int, max_scan: int = 3000) -> int:
    for r in range(start_row, start_row + max_scan):
        vals = [ws.cell(r, c).value for c in range(1, 6)]
        if all(v is None for v in vals):
            return r
    return start_row + max_scan

# 「入 ○○ (路線)  出 ○○ (路線)」から入駅・出駅を抜く
PAT_INOUT = re.compile(r"入\s*(?P<inn>.+?)\s*\(.+?\)\s*出\s*(?P<out>.+?)\s*\(.+?\)\s*$")
def extract_in_out(text: str):
    m = PAT_INOUT.search(str(text).strip())
    if not m:
        return None, None
    return m.group("inn").strip(), m.group("out").strip()

# 警告色（薄い赤）
ALERT_FILL = PatternFill(fill_type="solid", start_color="FFFFC7CE", end_color="FFFFC7CE")

# ============================================================
# アプリ本体
# ============================================================

st.title("交通費集計（まとめてドロップ → 増分追記 → D列だけ色付け）")

st.write(
    "使い方：**CSV と テンプレExcel（.xlsx）を2つまとめてドラッグ＆ドロップ** → 追記対象だけ転記 → ダウンロード\n\n"
    "仕様：\n"
    "- 移動区間(D列)＝CSVの「内容」そのまま\n"
    "- 備考は使わない\n"
    "- 訪問先・目的地(B列)：テンプレB5の文字が「出」側に含まれたら自宅、それ以外は同行\n"
    "- 交通手段(C列)：**電車（固定）**\n"
    "- 出力ファイル名：（JACOM・締め日MMDD）氏名（テンプレB4）\n"
    "- 既にテンプレにある行は触らず、重複しない分だけ追記\n"
    "- 前行の「出」駅 と 次行の「入」駅 が一致しなければ、移動区間(D列)だけ赤くする（固定）\n"
)

# まとめてドロップ（2ファイル）
files = st.file_uploader(
    "CSV と テンプレExcel（.xlsx）をまとめてドロップ（2ファイル）",
    type=["csv", "xlsx"],
    accept_multiple_files=True
)

csv_file = None
xlsx_file = None
if files:
    for f in files:
        name = f.name.lower()
        if name.endswith(".csv") and csv_file is None:
            csv_file = f
        elif name.endswith(".xlsx") and xlsx_file is None:
            xlsx_file = f

if not (csv_file and xlsx_file):
    st.info("CSV と テンプレExcel（.xlsx）を2つまとめてドロップしてください。")
    st.stop()

st.subheader("設定（基本は触らなくてOK）")
sheet_name_default = "【12月】交通費"
cell_name = st.text_input("氏名セル（B4）", value="B4")
cell_home = st.text_input("最寄駅セル（B5）", value="B5")
cell_close = st.text_input("締め日セル（例：G1）", value="G1")

# Excel読み込み
xlsx_bytes = BytesIO(xlsx_file.getvalue())
wb = openpyxl.load_workbook(xlsx_bytes)

sheetnames = wb.sheetnames
if sheet_name_default not in sheetnames:
    st.warning(f"シート「{sheet_name_default}」が見つかりません。候補から選んでください。")
    sheet_name = st.selectbox("転記先シート", sheetnames)
else:
    sheet_name = st.selectbox("転記先シート", sheetnames, index=sheetnames.index(sheet_name_default))

ws = wb[sheet_name]

# テンプレ情報
name_from_b4 = norm_str(ws[cell_name].value)
home_keyword = norm_str(ws[cell_home].value)
closing_date = get_date_from_cell(ws[cell_close].value)

if not name_from_b4:
    st.error("テンプレの氏名（B4）が空です。")
    st.stop()
if not closing_date:
    st.error("テンプレの締め日セルが日付として読めません。セル番地や値を確認してください。")
    st.stop()

st.info(f"テンプレ情報：氏名={name_from_b4} / 最寄駅(B5)={home_keyword} / 締め日={closing_date}")

# CSV読み込み
df = read_csv_safely(csv_file)

col_date = find_col(df, ["日付", "利用日", "日時", "日"])
col_content = find_col(df, ["内容", "経路", "明細", "区間"])
col_amount = find_col(df, ["金額", "運賃", "金額(円)", "利用金額", "支払金額"])

if not (col_date and col_content and col_amount):
    st.error(f"CSV列が足りません。必要: 日付/内容/金額。現在: {list(df.columns)}")
    st.stop()

# CSV → 正規化（自宅/同行判定）
rows = []
for _, r in df.iterrows():
    content = norm_str(r.get(col_content, ""))

    d = pd.to_datetime(r.get(col_date), errors="coerce")
    if pd.isna(d):
        continue
    d = d.date()

    amt = int(abs(pd.to_numeric(r.get(col_amount, 0), errors="coerce") or 0))

    dest = "同行"
    if home_keyword:
        inn, outt = extract_in_out(content)
        if outt and (home_keyword in outt):
            dest = "自宅"
        else:
            if ("出" in content) and (home_keyword in content[content.find("出"):]):
                dest = "自宅"

    rows.append({
        "日付": d,
        "訪問先・目的地": dest,
        "交通手段": "電車",          # ★電車固定
        "移動区間": content,
        "金額": amt
    })

new_df = pd.DataFrame(rows).sort_values(["日付", "移動区間"], kind="stable")

st.subheader("CSVプレビュー（上位20行）")
st.dataframe(new_df.head(20), use_container_width=True)

# 書き込み開始行を特定
header_row = find_header_row(ws, "日付")
if header_row is None:
    st.error("テンプレ内で「日付」ヘッダ行が見つかりません。")
    st.stop()
start_row = header_row + 1

# 既存キー（転記済み）を収集
existing_keys = collect_existing_keys(ws, start_row)

def make_key(row):
    return (row["日付"], row["訪問先・目的地"], row["交通手段"], row["移動区間"], int(row["金額"]))

new_df["__key__"] = new_df.apply(make_key, axis=1)
to_add = new_df[~new_df["__key__"].isin(existing_keys)].drop(columns=["__key__"])

st.subheader("判定結果")
st.write(f"CSV有効行数: {len(new_df)}")
st.write(f"重複スキップ: {len(new_df) - len(to_add)} 行")
st.write(f"追記対象: {len(to_add)} 行")

st.subheader("追記プレビュー（上位30行）")
st.dataframe(to_add.head(30), use_container_width=True)

# 追記開始行
first_empty = find_first_empty_row(ws, start_row)

# 直前行（既存の最終行）の「出」駅を拾う（色付け判定の起点）
prev_out_station = None
last_row = first_empty - 1
if last_row >= start_row:
    prev_route = ws.cell(last_row, 4).value  # D列
    _, prev_out_station = extract_in_out(prev_route)

# 追記 + 色付け（D列だけ固定）
date_fmt = 'm"月"d"日"'
rr = first_empty
flagged = 0

for _, row in to_add.iterrows():
    in_station, out_station = extract_in_out(row["移動区間"])
    mismatch = (
        prev_out_station is not None
        and in_station is not None
        and prev_out_station != in_station
    )

    c_date = ws.cell(rr, 1)
    c_date.value = row["日付"]
    c_date.number_format = date_fmt

    ws.cell(rr, 2).value = row["訪問先・目的地"]
    ws.cell(rr, 3).value = row["交通手段"]
    ws.cell(rr, 4).value = row["移動区間"]
    ws.cell(rr, 5).value = int(row["金額"])
    ws.cell(rr, 6).value = None
    ws.cell(rr, 7).value = None

    if mismatch:
        ws.cell(rr, 4).fill = ALERT_FILL  # ★D列だけ赤
        flagged += 1

    if out_station is not None:
        prev_out_station = out_station

    rr += 1

st.info(f"出→次の入 不一致で色付けした件数: {flagged}")

# 出力
mmdd = f"{closing_date.month:02d}{closing_date.day:02d}"
out_name = f"（JACOM・{mmdd}）{name_from_b4}.xlsx"

buf = BytesIO()
wb.save(buf)
buf.seek(0)

st.success(f"出力ファイル名：{out_name}")
st.download_button(
    "増分転記済みExcelをダウンロード",
    data=buf,
    file_name=out_name,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
